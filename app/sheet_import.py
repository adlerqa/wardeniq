"""Import-Sheet feature — parser + relevance scorer.

Ported from Node's `backend/src/services/testImportIntelligence.ts` (header alias
detection, row field derivation) and the algorithmic scoring formula in
`backend/src/services/testImport.service.ts`.

Accepts CSV / TSV / XLSX uploads, finds the header row (within first 15), maps
columns to a canonical schema, extracts rows, then scores each row's relevance
against a feature context. Output is one of:

  - 'matched'           → promote into the feature's test cases now
  - 'stored_for_later'  → keep in project-wide pool; user can pull later

Public API:

  parse_sheet(file_bytes, filename) -> list[ParsedRow]
  derive_row_fields(raw_row, header_map)
  build_feature_context(feature_doc) -> FeatureContext
  score_row(row, ctx) -> {score, action, hard_signals, breakdown}
  identity_hash(row) -> str           # canonical dedup key
  content_signature(rows) -> str       # whole-sheet fingerprint
  build_xlsx_template() -> bytes
  build_csv_template()  -> bytes
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from dataclasses import dataclass, field
from typing import Any

# ---------------------------------------------------------------- header aliases
# Ordered alias lists. The FIRST alias whose normalized form exactly matches a
# header cell wins. This matches Node's `HEADER_ALIASES` + `getColumnValue`
# behavior — substring matches like "test" inside "test id" were causing the
# Title field to greedily snap to the Test-ID column.
HEADER_ALIASES: dict[str, list[str]] = {
    # CRITICAL: "test id" / "case id" / "id" come BEFORE the title aliases
    # (and "test" / "name" / "testcase") so a "Test ID" column never gets
    # misclaimed as Title.
    "test_id":          ["test id", "case id", "tc id", "tcid", "key", "id"],
    "title":            ["title", "case title", "test case", "testcase",
                         "scenario", "test name", "name", "test"],
    "description":      ["description", "desc", "summary", "details",
                         "scenario description", "version", "document path"],
    "intent":           ["intent", "purpose", "goal", "objective"],
    "category":         ["category", "test type", "document type", "type"],
    "suite":            ["test suite", "suite"],
    "priority":         ["priority", "prio", "severity"],
    "endpoint":         ["endpoint", "endpoints", "path", "route", "url"],
    "method":           ["http method", "method", "verb"],
    "steps":            ["steps", "step", "procedure", "flow",
                         "test script step by step step",
                         "test script (step-by-step) - step"],
    "expected_result":  ["expected result", "expected outcome",
                         "expected result/behavior",
                         "expected result / behavior",
                         "expected", "outcome", "result",
                         "edge cases", "ui validations",
                         "test script (step-by-step) - expected result"],
    "module":           ["feature name", "feature area", "feature",
                         "module", "area", "domain", "component"],
    "tags":             ["labels", "tags", "tag", "label"],
    "preconditions":    ["preconditions", "precondition",
                         "preconditions/setup", "setup", "given"],
    "status":           ["status", "state"],
}

# Header normalization mirrors Node `normalizeHeaderKey`: lowercase, strip
# BOM, collapse anything that isn't [a-z0-9] to single spaces.
_HEADER_NORM_RE = re.compile(r"[^a-z0-9]+")


def _norm_header(s: str) -> str:
    s = (s or "").strip().lstrip("﻿").lower()
    return _HEADER_NORM_RE.sub(" ", s).strip()


def build_header_map(row_values: list[str]) -> dict[str, int]:
    """Map canonical field -> column index using EXACT normalized-key match.

    Iterates canonical fields in dict order (so test_id is claimed before title);
    inside each field, iterates its alias list in priority order. First exact
    match wins. Falls back to substring containment only when no exact match
    found AND the substring is the WHOLE normalized cell or a clear prefix —
    this preserves recall on slightly varied column names ("Steps (proc)") while
    still keeping "Test ID" out of the Title slot."""
    out: dict[str, int] = {}
    normalized = [_norm_header(v) for v in row_values]
    used_cols: set[int] = set()

    # Pass 1: exact equality
    for canonical, aliases in HEADER_ALIASES.items():
        if canonical in out:
            continue
        for alias in aliases:
            akey = _norm_header(alias)
            for i, cell in enumerate(normalized):
                if i in used_cols or not cell:
                    continue
                if cell == akey:
                    out[canonical] = i
                    used_cols.add(i)
                    break
            if canonical in out:
                break

    # Pass 2: prefix containment for unmatched fields (e.g. "Steps (numbered)")
    for canonical, aliases in HEADER_ALIASES.items():
        if canonical in out:
            continue
        for alias in aliases:
            akey = _norm_header(alias)
            for i, cell in enumerate(normalized):
                if i in used_cols or not cell:
                    continue
                # Whole-word prefix OR whole-word contained.
                if cell.startswith(akey + " ") or cell.endswith(" " + akey) \
                        or (" " + akey + " ") in (" " + cell + " "):
                    out[canonical] = i
                    used_cols.add(i)
                    break
            if canonical in out:
                break
    return out


# ---------------------------------------------------------------- header detect
_HEADER_KEYWORDS = {"title", "name", "scenario", "steps", "expected", "test",
                     "case", "description", "priority", "category", "endpoint",
                     "method", "module"}


def is_likely_header_row(row_values: list[str]) -> bool:
    """Heuristic: row qualifies as header when ≥2 cells normalize to a known
    canonical alias keyword. Matches Node's `isLikelyHeaderRow`."""
    hits = 0
    for v in row_values:
        nv = _norm_header(v)
        if not nv:
            continue
        for kw in _HEADER_KEYWORDS:
            if kw in nv:
                hits += 1
                break
        if hits >= 2:
            return True
    return False


# ---------------------------------------------------------------- row extraction
@dataclass
class ParsedRow:
    """One extracted row from the sheet, normalized."""
    sheet: str
    row_number: int
    title: str = ""
    description: str = ""
    intent: str = ""
    category: str | None = None
    suite: str = ""
    priority: str = "Mid"
    endpoint: str = ""
    method: str = ""
    steps: list[str] = field(default_factory=list)
    expected_result: str = ""
    module: str = ""
    tags: list[str] = field(default_factory=list)
    preconditions: str = ""
    test_id: str = ""
    status: str = ""
    raw: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict:
        return {
            "title": self.title, "description": self.description,
            "intent": self.intent, "category": self.category, "suite": self.suite,
            "priority": self.priority, "endpoint": self.endpoint,
            "method": self.method, "steps": self.steps,
            "expected_result": self.expected_result, "module": self.module,
            "tags": self.tags, "preconditions": self.preconditions,
            "test_id": self.test_id, "status": self.status,
            "sheet": self.sheet, "row_number": self.row_number,
        }


_STEP_SPLIT = re.compile(r"\n+|;\s*|\|\s*|\d+\.\s+", re.MULTILINE)
_TAG_SPLIT = re.compile(r"[,;\|]")
_ROUTE_RE = re.compile(
    r"\b(?:GET|POST|PUT|PATCH|DELETE|HEAD|OPTIONS)\s+(/[A-Za-z0-9_./:{}?=&%-]+)"
    r"|(/[A-Za-z0-9_./:{}-]+(?:/[A-Za-z0-9_./:{}-]+)+)",
    re.IGNORECASE,
)
_VERSION_NOISE_RE = re.compile(r"\b(?:version|v)\s*\d+(?:\.\d+)*\b", re.IGNORECASE)
_PRIORITY_NORM = {
    "1": "High", "high": "High", "h": "High", "p0": "High", "p1": "High",
    "critical": "High",
    "2": "Mid", "mid": "Mid", "medium": "Mid", "m": "Mid", "p2": "Mid", "normal": "Mid",
    "3": "Low", "low": "Low", "l": "Low", "p3": "Low", "p4": "Low",
}


def _norm_priority(s: str) -> str:
    if not s:
        return "Mid"
    return _PRIORITY_NORM.get(s.strip().lower(), "Mid")


def _route_fragments(text: str) -> list[str]:
    out = []
    for match in _ROUTE_RE.finditer(text or ""):
        value = match.group(0).strip()
        if value and value not in out:
            out.append(value)
    return out


def _looks_like_route_dump(text: str) -> bool:
    routes = _route_fragments(text)
    return len(routes) >= 2 or (len(routes) == 1 and len((text or "").split()) <= 8)


def _clean_step_text(text: str) -> str:
    cleaned = re.sub(r"^\s*steps?\s*:\s*", "", text or "", flags=re.IGNORECASE).strip()
    cleaned = _VERSION_NOISE_RE.sub("", cleaned).strip(" -·:|")
    # Empty fragments (e.g. the gap left by splitting "1. foo\n2. bar") must be
    # dropped, NOT turned into a placeholder step — otherwise every numbered list
    # gets padded with bogus "Validate documented API behavior" entries.
    if not cleaned:
        return ""
    if _looks_like_route_dump(cleaned):
        return "Validate documented API behavior"
    return cleaned


def _clean_expected_text(text: str) -> str:
    cleaned = (text or "").strip()
    if _looks_like_route_dump(cleaned):
        return "The documented API behavior is validated without exposing raw endpoint dumps in the test steps."
    return cleaned


def derive_row_fields(raw_row: list[str], header_map: dict[str, int],
                       sheet: str, row_number: int) -> ParsedRow | None:
    """Pluck canonical fields out of a raw cell list. Returns None when the row
    has nothing usable (no title/description/steps)."""
    def cell(name: str) -> str:
        idx = header_map.get(name, -1)
        if idx < 0 or idx >= len(raw_row):
            return ""
        v = raw_row[idx]
        if v is None:
            return ""
        return str(v).strip()

    title = cell("title")
    description = cell("description")
    intent = cell("intent")
    steps_raw = cell("steps")
    expected = cell("expected_result")
    if not (title or description or steps_raw):
        return None

    steps = []
    if steps_raw:
        for part in _STEP_SPLIT.split(steps_raw):
            p = _clean_step_text(part)
            if p:
                steps.append(p)

    tags_raw = cell("tags")
    tags = [t.strip() for t in _TAG_SPLIT.split(tags_raw) if t.strip()]

    return ParsedRow(
        sheet=sheet, row_number=row_number,
        title=title, description=description, intent=intent,
        category=(cell("category") or None),
        suite=cell("suite"),
        priority=_norm_priority(cell("priority")),
        endpoint=cell("endpoint"), method=cell("method").upper(),
        steps=steps, expected_result=_clean_expected_text(expected),
        module=cell("module"), tags=tags,
        preconditions=cell("preconditions"),
        test_id=cell("test_id"), status=cell("status"),
        raw={"original": raw_row},
    )


# ---------------------------------------------------------------- file parsing
def _decode_text(file_bytes: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return file_bytes.decode(enc)
        except Exception:  # noqa: BLE001
            continue
    return file_bytes.decode("utf-8", errors="replace")


def _looks_like_bug_sheet(sheet_name: str) -> bool:
    n = (sheet_name or "").lower()
    return "bug" in n or "defect" in n or "issue" in n


def parse_sheet(file_bytes: bytes, filename: str) -> list[ParsedRow]:
    """Read CSV / TSV / XLSX from bytes and return ParsedRow list across all
    sheets. Skips sheets named like 'Bugs'/'Defects' (matches Node)."""
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _parse_xlsx(file_bytes)
    if name.endswith(".tsv"):
        return _parse_csv(file_bytes, delimiter="\t")
    return _parse_csv(file_bytes, delimiter=",")


def raw_tables(file_bytes: bytes, filename: str) -> list[tuple[str, list[list[str]]]]:
    """Return (sheet_name, raw_rows[]) BEFORE structured parsing. Used to feed
    the LLM QA-relevance classifier with the same cell content Node sees, so
    the prompt judges the actual upload rather than already-cleaned rows."""
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        try:
            import openpyxl
        except ImportError:
            return []
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True,
                                     read_only=True)
        out = []
        for sheet in wb.worksheets:
            rows = []
            for row in sheet.iter_rows(values_only=True):
                cells = [("" if c is None else str(c)) for c in row]
                if any(c.strip() for c in cells):
                    rows.append(cells)
                if len(rows) >= 30:
                    break
            out.append((sheet.title or "Sheet1", rows))
        return out
    if name.endswith(".tsv"):
        text = _decode_text(file_bytes)
        reader = csv.reader(io.StringIO(text), delimiter="\t")
        rows = [list(r) for r in reader if any((c or "").strip() for c in r)][:30]
        return [("Sheet1", rows)]
    text = _decode_text(file_bytes)
    reader = csv.reader(io.StringIO(text))
    rows = [list(r) for r in reader if any((c or "").strip() for c in r)][:30]
    return [("Sheet1", rows)]


# ---------------------------------------------------------------- hardcoded fallback
_QA_KEYWORDS = re.compile(
    r"\b(test\s?case|test\s?id|test\s?name|test\s?suite|test\s?type|"
    r"scenario|precondition|expected\s?result|expected\s?outcome|"
    r"steps?|priority|severity|module|story|requirement|defect|bug)\b",
    re.IGNORECASE,
)


def looks_like_qa_sheet_heuristic(tables: list[tuple[str, list[list[str]]]]) -> bool:
    """Hardcoded fallback when the LLM classifier fails or is unconfigured.

    Returns True when the raw cells contain enough QA-vocabulary signal that
    we'd accept the sheet without LLM help. Used both as a pre-LLM short-circuit
    and as the safety net on LLM error."""
    if not tables:
        return False
    hits = 0
    cells_seen = 0
    for _, rows in tables:
        for r in rows[:20]:
            for c in r:
                cells_seen += 1
                if _QA_KEYWORDS.search(c or ""):
                    hits += 1
        if cells_seen >= 200:
            break
    return hits >= 2


def _parse_csv(file_bytes: bytes, delimiter: str = ",") -> list[ParsedRow]:
    # Binary content (e.g. a real .xlsx/.pdf renamed to .csv) decodes to text
    # full of NULs and stray newlines that blow up csv.reader. Detect that and
    # raise a readable message instead of a cryptic csv.Error.
    if b"\x00" in file_bytes[:4096]:
        raise ValueError(
            "This doesn't look like a text/CSV file. Please upload a CSV, "
            "TSV, or XLSX spreadsheet.")
    text = _decode_text(file_bytes).replace("\x00", "")
    try:
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows = [list(r) for r in reader]
    except csv.Error as e:
        raise ValueError(
            f"Could not read this file as a spreadsheet ({e}). Please upload "
            "a valid CSV, TSV, or XLSX file.") from e
    return _process_table(rows, "Sheet1")


def _parse_xlsx(file_bytes: bytes) -> list[ParsedRow]:
    """Parse XLSX with merged-cell anchor propagation. When a sheet has a
    merge range (e.g. Title spans rows 5-7), every cell in the range is
    populated with the anchor value — otherwise openpyxl returns None for the
    non-anchor cells and continuation rows lose their title."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed — add it to requirements.txt")
    # NB: read_only=False so `merged_cells` is populated.
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True,
                                     read_only=False)
    except Exception as e:  # noqa: BLE001 — BadZipFile, InvalidFileException, …
        raise ValueError(
            "This file isn't a valid .xlsx workbook (it may be corrupt, a "
            "legacy .xls, or a different file type renamed to .xlsx). Please "
            "re-export and upload a valid .xlsx or CSV file.") from e
    out: list[ParsedRow] = []
    for sheet in wb.worksheets:
        if _looks_like_bug_sheet(sheet.title):
            continue
        # Build a (row, col) -> anchor-value map for every merged range.
        merge_lookup: dict[tuple[int, int], object] = {}
        for mr in list(sheet.merged_cells.ranges):
            anchor = sheet.cell(row=mr.min_row, column=mr.min_col).value
            if anchor is None or str(anchor).strip() == "":
                continue
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    merge_lookup[(r, c)] = anchor
        rows: list[list[str]] = []
        max_col = sheet.max_column or 0
        for ri in range(1, (sheet.max_row or 0) + 1):
            cells = []
            for ci in range(1, max_col + 1):
                v = sheet.cell(row=ri, column=ci).value
                if v is None and (ri, ci) in merge_lookup:
                    v = merge_lookup[(ri, ci)]
                cells.append("" if v is None else str(v))
            # Drop rows that are entirely whitespace after propagation.
            if any(c.strip() for c in cells):
                rows.append(cells)
        out.extend(_process_table(rows, sheet.title or "Sheet1"))
    return out


def _process_table(rows: list[list[str]], sheet_name: str) -> list[ParsedRow]:
    """Find the header row within the first 15 rows, parse rows, then group
    consecutive rows that share the same title/key into one test case
    (Jira/Zephyr/grouped-export shape: title on the first row, more steps +
    expected-results on the rows immediately following)."""
    if not rows:
        return []
    header_row_index = -1
    for i, r in enumerate(rows[:15]):
        if is_likely_header_row(r):
            header_row_index = i
            break
    if header_row_index < 0:
        for i, r in enumerate(rows[:5]):
            if any((c or "").strip() for c in r):
                header_row_index = i
                break
    if header_row_index < 0:
        return []
    header_map = build_header_map(rows[header_row_index])
    if not header_map:
        return []

    # ---- pass 1: derive raw per-row records (some may have no title) -------
    raw_records: list[tuple[int, list[str], ParsedRow | None]] = []
    for i in range(header_row_index + 1, len(rows)):
        raw = rows[i]
        joined = " ".join((c or "").strip() for c in raw).strip()
        if not joined:
            continue
        if "[example]" in joined.lower():
            continue
        rec = derive_row_fields_loose(raw, header_map, sheet_name, i + 1)
        raw_records.append((i + 1, raw, rec))

    # ---- pass 2: group consecutive rows that share title or test_id --------
    grouped: list[ParsedRow] = []
    for row_num, raw, rec in raw_records:
        if rec is None:
            continue
        if grouped and _is_continuation_of(rec, grouped[-1]):
            _merge_into(grouped[-1], rec)
        else:
            grouped.append(rec)

    # ---- pass 3: drop rows that ended up empty after grouping --------------
    final: list[ParsedRow] = []
    for r in grouped:
        if not (r.title or r.description or r.steps or r.expected_result):
            continue
        # If row still has no title but has description, hoist description.
        if not r.title and r.description:
            r.title = r.description if len(r.description) <= 200 else r.description[:200]
        if not r.title and r.steps:
            r.title = r.steps[0][:200]
        if not r.title:
            continue
        # Trim oversized fields.
        if len(r.title) > 240:
            r.title = r.title[:240]
        final.append(r)
    return final


def derive_row_fields_loose(raw_row: list[str], header_map: dict[str, int],
                              sheet: str, row_number: int) -> ParsedRow | None:
    """Like derive_row_fields, but RETAINS continuation rows that lack a title
    so the grouping pass can merge their steps/expected into the parent."""
    def cell(name: str) -> str:
        idx = header_map.get(name, -1)
        if idx < 0 or idx >= len(raw_row):
            return ""
        v = raw_row[idx]
        if v is None:
            return ""
        return str(v).strip()

    title = cell("title")
    description = cell("description")
    intent = cell("intent")
    steps_raw = cell("steps")
    expected = cell("expected_result")
    test_id = cell("test_id")
    endpoint = cell("endpoint")
    method = cell("method").upper()
    module = cell("module")
    category = cell("category") or None
    # A "document catalog" is a sheet with NO title column at all, where every
    # row must synthesize a title from its module/endpoint. When the sheet DOES
    # have a title column (Zephyr/Jira grouped exports), title-less rows are
    # continuation rows — we must NOT synthesize a fake title for them, or the
    # grouping pass can't merge their steps into the parent test case.
    is_document_catalog = (
        "title" not in header_map
        and "module" in header_map
        and ("endpoint" in header_map or "expected_result" in header_map)
        and ("category" in header_map or "description" in header_map)
    )
    if is_document_catalog:
        has_qa_evidence = bool(endpoint or expected or steps_raw)
        if not has_qa_evidence:
            return None
        if not title:
            subject = module or endpoint or "documented behavior"
            qualifier = category or "QA"
            title = f"Validate {subject} {qualifier} behavior".strip()
        if not description:
            description = (
                f"Reusable QA evidence imported for {module or title}."
            )
        if not steps_raw:
            steps_raw = "Validate documented API behavior"
        if not expected:
            expected = "Documented behavior is validated successfully."
    # NEW: retain row even when title is empty — grouping merges later.
    if not (title or description or steps_raw or expected or test_id):
        return None

    steps = []
    if steps_raw:
        for part in _STEP_SPLIT.split(steps_raw):
            p = _clean_step_text(part)
            if p:
                steps.append(p)

    tags_raw = cell("tags")
    tags = [t.strip() for t in _TAG_SPLIT.split(tags_raw) if t.strip()]

    return ParsedRow(
        sheet=sheet, row_number=row_number,
        title=title, description=description, intent=intent,
        category=category,
        suite=cell("suite"),
        priority=_norm_priority(cell("priority")),
        endpoint=endpoint, method=method,
        steps=steps, expected_result=_clean_expected_text(expected),
        module=module, tags=tags,
        preconditions=cell("preconditions"),
        test_id=test_id, status=cell("status"),
        raw={"original": raw_row},
    )


def _is_continuation_of(rec: ParsedRow, parent: ParsedRow) -> bool:
    """True if `rec` should be merged into `parent` as additional steps/expected.

    Continuation when:
      - rec has NO title and NO test_id, but has steps or expected
      - OR rec has same test_id as parent (and parent.test_id is non-empty)
      - OR rec has identical title to parent
    """
    # When both rows carry an explicit ID, they belong together ONLY if the IDs
    # match — two distinct IDs are always separate cases, even with equal titles.
    if rec.test_id and parent.test_id:
        return rec.test_id == parent.test_id
    if rec.title and parent.title and rec.title == parent.title:
        return True
    if not rec.title and not rec.test_id and (rec.steps or rec.expected_result
                                                or rec.description):
        return True
    return False


def _merge_into(parent: ParsedRow, child: ParsedRow) -> None:
    """Combine child's steps + expected + description into parent."""
    for s in child.steps:
        if s and s not in parent.steps:
            parent.steps.append(s)
    if child.expected_result:
        if parent.expected_result:
            if child.expected_result not in parent.expected_result:
                parent.expected_result = (parent.expected_result + " | "
                                            + child.expected_result)[:1500]
        else:
            parent.expected_result = child.expected_result[:1500]
    if child.description and child.description not in parent.description:
        if parent.description:
            parent.description = (parent.description + " "
                                   + child.description)[:1500]
        else:
            parent.description = child.description[:1500]
    for t in child.tags:
        if t not in parent.tags:
            parent.tags.append(t)


# Back-compat alias — earlier callers expect the strict version. Now defers to
# the same loose extractor so external code keeps working.
def derive_row_fields(raw_row: list[str], header_map: dict[str, int],
                       sheet: str, row_number: int) -> ParsedRow | None:
    return derive_row_fields_loose(raw_row, header_map, sheet, row_number)


# ---------------------------------------------------------------- identity & sig
_NORM_RE = re.compile(r"\s+")


def _normalize(s: str) -> str:
    return _NORM_RE.sub(" ", (s or "").strip().lower())


def identity_hash(row: ParsedRow | dict) -> str:
    """Canonical hash so the same row across multiple sheets dedupes to one
    canonical row in the project pool."""
    if isinstance(row, ParsedRow):
        d = row.to_dict()
    else:
        d = row
    parts = [
        _normalize(d.get("title", "")),
        _normalize(d.get("endpoint", "")),
        _normalize(d.get("method", "")),
        "|".join(_normalize(s) for s in (d.get("steps") or [])),
        _normalize(d.get("expected_result", "")),
    ]
    return hashlib.sha256("\x1f".join(parts).encode("utf-8")).hexdigest()


def content_signature(rows: list[ParsedRow]) -> str:
    """SHA-256 fingerprint of the whole sheet — used to skip re-processing the
    exact same upload."""
    h = hashlib.sha256()
    h.update(f"rows={len(rows)}\n".encode())
    for r in rows:
        h.update((r.sheet or "").encode())
        h.update(b"\x1f")
        h.update(identity_hash(r).encode())
        h.update(b"\x1e")
    return h.hexdigest()


def normalize_imported_payload_shape(payload: dict[str, Any]) -> dict[str, Any]:
    """Repair legacy imported-row payloads before display or promotion."""
    data = dict(payload or {})
    raw_steps = data.get("steps") or []
    if isinstance(raw_steps, str):
        raw_steps = [p for p in _STEP_SPLIT.split(raw_steps) if p.strip()]
    cleaned_steps = []
    for step in raw_steps:
        if isinstance(step, dict):
            content = _clean_step_text(str(step.get("content") or step.get("step") or ""))
            expected = _clean_expected_text(str(step.get("expectedResult")
                                               or step.get("expected_result") or ""))
            if expected:
                cleaned_steps.append({"content": content, "expectedResult": expected})
            else:
                cleaned_steps.append(content)
        else:
            cleaned_steps.append(_clean_step_text(str(step)))
    data["steps"] = [s for s in cleaned_steps if s]
    expected = data.get("expected_result")
    if isinstance(expected, dict):
        expected = expected.get("summary") or expected.get("text") or ""
    data["expected_result"] = _clean_expected_text(str(expected or ""))
    if not data.get("title") and data["steps"]:
        first = data["steps"][0]
        data["title"] = (first.get("content") if isinstance(first, dict) else str(first))[:200]
    if not data.get("priority"):
        data["priority"] = "Mid"
    return data


# ---------------------------------------------------------------- feature ctx
_INTENT_SYNONYMS = [
    (re.compile(r"\bsign[\s-]?in\b|\blog[\s-]?in\b|\bauthenticate\b"), " auth "),
    (re.compile(r"\bsign[\s-]?up\b|\bregister\b|\bonboard\b"), " signup "),
    (re.compile(r"\b(?:create|add|insert|new)\b"), " create "),
    (re.compile(r"\b(?:update|edit|modify|change|patch)\b"), " update "),
    (re.compile(r"\b(?:delete|remove|destroy|drop)\b"), " delete "),
    (re.compile(r"\b(?:get|fetch|retrieve|view|show|list)\b"), " read "),
    (re.compile(r"\b(?:upload|attach)\b"), " upload "),
    (re.compile(r"\b(?:download|export)\b"), " download "),
    (re.compile(r"\b(?:invite|share)\b"), " invite "),
    (re.compile(r"\b(?:notify|notification|push|email)\b"), " notify "),
    (re.compile(r"\bpayment|charge|invoice|billing\b"), " payment "),
    (re.compile(r"\b(?:cancel|refund)\b"), " cancel "),
    (re.compile(r"\b(?:validate|verify|check)\b"), " validate "),
]


def normalize_intent(s: str) -> str:
    t = " " + (s or "").lower() + " "
    for rx, repl in _INTENT_SYNONYMS:
        t = rx.sub(repl, t)
    return _NORM_RE.sub(" ", t).strip()


def extract_intent_concepts(s: str) -> set[str]:
    nt = normalize_intent(s)
    return {tok for tok in nt.split() if len(tok) > 2}


_STOP = {"a","an","the","and","or","of","for","to","in","on","with","as","is",
          "are","be","by","this","that","it","its","at","from","into","onto",
          "when","then","given","should","must","will","shall","may","might",
          "test","case","verify","check","ensure","via","such","using","over",
          "across","also","any","all","each","other","because","but","so","do",
          "does","done","not","no","yes"}
_TOKEN = re.compile(r"[a-z][a-z0-9_]{1,}")


def tokenize(s: str) -> set[str]:
    """Lower-case, drop stop-words, min length 2 — mirrors Node `tokenize`."""
    return {t for t in _TOKEN.findall((s or "").lower())
            if t not in _STOP and len(t) > 1}


@dataclass
class FeatureContext:
    feature_id: str
    name: str
    description: str
    intent: str
    intent_norm: str
    intent_concepts: set[str]
    name_tokens: set[str]
    feature_tokens: set[str]      # all tokens from name+desc+intent
    anchor_tokens: set[str]        # noun-like tokens from name (≥4 chars)
    endpoints: set[str]
    methods: set[str]
    route_families: set[str]       # e.g. "/users", "/orders"


def build_feature_context(feature: dict) -> FeatureContext:
    name = feature.get("name", "") or ""
    desc = feature.get("description") or feature.get("summary", "") or ""
    text_for_intent = name + " " + desc
    intent_norm = normalize_intent(text_for_intent)
    intent_concepts = extract_intent_concepts(text_for_intent)
    name_tokens = tokenize(name)
    feature_tokens = tokenize(name + " " + desc + " " + (feature.get("text") or "")[:2000])
    anchor_tokens = {t for t in name_tokens if len(t) >= 4}
    raw_api = feature.get("raw_api_spec") or []
    endpoints: set[str] = set()
    methods: set[str] = set()
    route_families: set[str] = set()
    if isinstance(raw_api, list):
        for ep in raw_api:
            if not isinstance(ep, dict):
                continue
            path = (ep.get("path") or "").lower().strip()
            method = (ep.get("method") or "").upper().strip()
            if path:
                endpoints.add(path)
                family = "/" + path.lstrip("/").split("/")[0]
                if family != "/":
                    route_families.add(family)
            if method:
                methods.add(method)
    return FeatureContext(
        feature_id=str(feature.get("id") or feature.get("_id") or ""),
        name=name, description=desc, intent=text_for_intent,
        intent_norm=intent_norm, intent_concepts=intent_concepts,
        name_tokens=name_tokens, feature_tokens=feature_tokens,
        anchor_tokens=anchor_tokens, endpoints=endpoints,
        methods=methods, route_families=route_families,
    )


# ---------------------------------------------------------------- scoring
def _jaccard(a: set, b: set) -> float:
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


def _overlap_ratio(a: set, b: set) -> float:
    if not b:
        return 0.0
    return len(a & b) / len(b)


def _endpoint_match(row_ep: str, ctx_endpoints: set[str]) -> float:
    if not row_ep or not ctx_endpoints:
        return 0.0
    rp = row_ep.lower().strip()
    for ep in ctx_endpoints:
        # exact / prefix
        if rp == ep or rp.startswith(ep) or ep.startswith(rp):
            return 1.0
        # share top-level family
        if rp.split("/")[1:2] == ep.split("/")[1:2]:
            return 0.5
    return 0.0


def _route_family_match(row_ep: str, families: set[str]) -> float:
    if not row_ep or not families:
        return 0.0
    rp = row_ep.lower().strip()
    if not rp.startswith("/"):
        rp = "/" + rp.lstrip("/")
    top = "/" + rp.lstrip("/").split("/")[0]
    return 1.0 if top in families else 0.0


def _completeness(row: ParsedRow) -> int:
    """Count of filled canonical fields (max 7) — used as a small recall bonus."""
    c = 0
    if row.title: c += 1
    if row.description: c += 1
    if row.steps: c += 1
    if row.expected_result: c += 1
    if row.endpoint: c += 1
    if row.method: c += 1
    if row.category: c += 1
    return c


def _category_alignment(row: ParsedRow, ctx: FeatureContext) -> float:
    """Modest bonus when category text mentions tokens from feature name."""
    cat = (row.category or "").lower()
    if not cat:
        return 0.0
    cat_tokens = tokenize(cat)
    if cat_tokens & ctx.name_tokens:
        return 1.0
    return 0.0


def _testcase_likelihood(row: ParsedRow) -> float:
    """Cheap signal: does the row LOOK like a test case (has steps + expected)?"""
    if row.steps and row.expected_result:
        return 1.0
    if row.steps or row.expected_result:
        return 0.6
    return 0.0


def _alien_penalty(row: ParsedRow, ctx: FeatureContext) -> float:
    """Penalty when row text is dense with tokens completely foreign to the
    feature — keeps obvious garbage rows out."""
    title_tokens = tokenize(row.title)
    if not title_tokens or not ctx.feature_tokens:
        return 0.0
    alien = title_tokens - ctx.feature_tokens
    if not alien:
        return 0.0
    ratio = len(alien) / max(1, len(title_tokens))
    return 0.25 if ratio >= 0.85 else 0.0


def _domain_tokens(row: ParsedRow) -> set[str]:
    """Tokens we treat as domain-meaningful — used for the min-score floor."""
    s = " ".join([row.title, row.endpoint, row.module, row.suite,
                  row.expected_result])
    return tokenize(s)


@dataclass
class ScoreResult:
    score: float
    action: str
    anchor_hits: int
    route_family: float
    endpoint_score: float
    intent_similarity: float
    intent_concept_overlap: float
    feature_name_overlap: float
    breakdown: dict


def score_row(row: ParsedRow, ctx: FeatureContext) -> ScoreResult:
    """Algorithmic relevance scorer — full Node port of `scoreRowRelevance`."""
    row_text = " ".join([
        row.title, row.description, row.intent, row.expected_result,
        " ".join(row.steps), row.module, row.suite,
    ])
    row_tokens = tokenize(row_text)
    title_tokens = tokenize(row.title)
    intent_tokens = tokenize(row.intent)

    title_overlap = _jaccard(title_tokens, ctx.name_tokens)
    row_overlap = _jaccard(row_tokens, ctx.feature_tokens)
    intent_overlap = _jaccard(intent_tokens, ctx.feature_tokens)
    intent_concept_overlap = _jaccard(extract_intent_concepts(row_text),
                                      ctx.intent_concepts)
    intent_similarity = _jaccard(tokenize(normalize_intent(row_text)),
                                  tokenize(ctx.intent_norm))
    anchor_overlap = _overlap_ratio(row_tokens, ctx.anchor_tokens)
    anchor_hits = len(row_tokens & ctx.anchor_tokens)
    feature_token_overlap = _jaccard(row_tokens, ctx.feature_tokens)
    feature_name_overlap = _overlap_ratio(row_tokens, ctx.name_tokens)
    route_family = _route_family_match(row.endpoint, ctx.route_families)
    testcase_likelihood = _testcase_likelihood(row)
    specificity = 1.0 if (row.endpoint or len(row.steps) >= 2) else 0.0
    category_alignment = _category_alignment(row, ctx)
    endpoint_score = _endpoint_match(row.endpoint, ctx.endpoints)
    completeness = _completeness(row)
    alien_penalty = _alien_penalty(row, ctx)

    # Node's weighted formula, verbatim:
    raw_score = (
        title_overlap * 0.03 +
        row_overlap * 0.05 +
        intent_overlap * 0.08 +
        intent_concept_overlap * 0.28 +
        intent_similarity * 0.24 +
        anchor_overlap * 0.18 +
        feature_token_overlap * 0.04 +
        feature_name_overlap * 0.02 +
        route_family * 0.16 +
        testcase_likelihood * 0.1 +
        specificity * 0.7 +
        category_alignment * 0.85 +
        endpoint_score * 0.9 +
        (completeness / 7.0) * 0.06 -
        alien_penalty
    )
    domain_tokens = _domain_tokens(row)
    min_score = 0.11 if len(domain_tokens) >= 2 else 0.0
    score = max(min_score, min(raw_score, 1.0))

    # Hard-signal AND-gate. Must clear 0.45 AND have at least one hard signal.
    hard_signal = (
        anchor_hits >= 2
        or route_family >= 0.15
        or endpoint_score >= 0.2
        or feature_name_overlap >= 0.2
        or intent_similarity >= 0.2
        or intent_concept_overlap >= 0.18
    )
    action = "matched" if (score >= 0.45 and hard_signal) else "stored_for_later"

    breakdown = {
        "title_overlap": round(title_overlap, 3),
        "row_overlap": round(row_overlap, 3),
        "intent_concept_overlap": round(intent_concept_overlap, 3),
        "intent_similarity": round(intent_similarity, 3),
        "anchor_overlap": round(anchor_overlap, 3),
        "route_family": round(route_family, 3),
        "endpoint_score": round(endpoint_score, 3),
        "category_alignment": round(category_alignment, 3),
        "specificity": specificity,
        "completeness": completeness,
        "alien_penalty": alien_penalty,
        "raw_score": round(raw_score, 3),
    }
    return ScoreResult(
        score=round(score, 3), action=action, anchor_hits=anchor_hits,
        route_family=route_family, endpoint_score=endpoint_score,
        intent_similarity=intent_similarity,
        intent_concept_overlap=intent_concept_overlap,
        feature_name_overlap=feature_name_overlap, breakdown=breakdown,
    )


# ---------------------------------------------------------------- templates
TEMPLATE_COLUMNS = [
    ("Title", 25),
    ("Description", 35),
    ("Intent", 25),
    ("Category", 15),
    ("Priority", 12),
    ("Endpoint", 20),
    ("Method", 10),
    ("Module", 15),
    ("Suite", 12),
    ("Steps", 35),
    ("Testcase Expected Result", 35),
]

TEMPLATE_EXAMPLE = [
    "[EXAMPLE] User Login with valid OTP",
    "Verify that an active user can successfully log in using a valid 6-digit OTP.",
    "Verify successful authentication using OTP.",
    "api_tests",
    "High",
    "/api/v1/auth/login",
    "POST",
    "Authentication",
    "Smoke",
    ("1. Send OTP to registered phone number\n"
     "2. Enter the received 6-digit OTP in the input field\n"
     "3. Click the Submit button"),
    "200 OK. User is redirected to the dashboard with a valid session.",
]


def build_csv_template() -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([c[0] for c in TEMPLATE_COLUMNS])
    w.writerow(TEMPLATE_EXAMPLE)
    return buf.getvalue().encode("utf-8")


# ---------------------------------------------------------------- LLM passes
# Mirrors Node's `isQARelatedSpreadsheet` (testImportIntelligence.ts:2303)
# and `polishImportedRowsWithAI` (testImport.service.ts:3553).

_QA_CLASSIFIER_SYS = (
    "You are a software quality assurance (QA) and test case catalog classifier. "
    "Respond with a single JSON object only."
)


def classify_sheet_is_qa(llm, tables_preview: str, feature_name: str,
                          feature_description: str) -> tuple[bool, str]:
    """Decide if the sheet is QA-related. Returns (is_qa, reason).
    Defaults to True on LLM failure (don't block imports on network blips)."""
    prompt = (
        "Analyze the structure and sample rows of an uploaded spreadsheet to "
        "determine if it is a valid software QA sheet.\n\n"
        "Valid QA sheets include:\n"
        "- Test case repositories / test case lists / test steps\n"
        "- Test plans / test scope / test strategy documents\n"
        "- QA checklists / manual test logs\n"
        "- Software bug log / defect tracking sheets\n"
        "- User story / requirement validation checklists\n\n"
        "Invalid/Unrelated sheets include:\n"
        "- Financial ledgers, sales logs, invoice tracking, budgets\n"
        "- General corporate lists, employee rosters, organization charts\n"
        "- Grocery lists, personal checklists, non-technical documents\n"
        "- Raw CSV code exports, system databases, unrelated data\n\n"
        f"Feature Name: {feature_name}\n"
        f"Feature Description: {feature_description or 'N/A'}\n\n"
        f"Spreadsheet preview:\n{tables_preview}\n\n"
        'Return JSON: {"isQARelated": true|false, '
        '"reason": "<one sentence>"}.'
    )
    try:
        data = llm.chat_json(_QA_CLASSIFIER_SYS, prompt, temperature=0.1,
                              max_tokens=400, retries=0)
    except Exception as e:  # noqa: BLE001
        return True, f"classifier failed: {e}"
    is_qa = data.get("isQARelated", True) is not False
    return is_qa, str(data.get("reason", "ok"))[:300]


def build_tables_preview(rows_by_sheet: list[tuple[str, list[list[str]]]]) -> str:
    """Build a compact preview of the upload (up to ~3000 chars)."""
    out = []
    used = 0
    for sheet_name, rows in rows_by_sheet:
        non_empty = [r for r in rows
                     if any((c or "").strip() for c in r)]
        first = "\n".join(" | ".join(r) for r in non_empty[:8])
        block = (f'Worksheet Name: "{sheet_name}" ({len(non_empty)} rows)\n'
                 f'Sample Rows:\n{first}\n\n')
        out.append(block)
        used += len(block)
        if used > 3000:
            break
    return "".join(out)


_POLISH_SYS = (
    "You are a senior QA intelligence engine. You polish QA rows imported "
    "from a user spreadsheet into PROJECT-WIDE reusable testcase memory. "
    "Respond with a single JSON object only."
)


def polish_rows_batch(llm, rows: list[ParsedRow],
                       feature_name: str, feature_description: str) -> dict:
    """One LLM call polishes up to ~8 rows. Returns the parsed JSON or {}.

    Each polished row carries `is_testcase` so metadata/header rows can be
    dropped, plus normalized title/desc/intent/category/priority/endpoint/
    method/steps/expected_result/tags."""
    if not rows:
        return {"polishedRows": []}
    rows_data = []
    for r in rows:
        rows_data.append({
            "rowNumber": r.row_number,
            "title": r.title,
            "description": r.description,
            "intent": r.intent,
            "category": r.category,
            "priority": r.priority,
            "endpoint": r.endpoint,
            "method": r.method,
            "steps": r.steps,
            "expected_result": r.expected_result,
            "tags": r.tags,
        })
    prompt = (
        "RULES:\n"
        "1. Preserve user-provided values. Don't rewrite titles/descriptions/"
        "steps unless fixing obvious formatting issues.\n"
        "2. Don't prepend feature names, version labels, or document type "
        "references to step content. Steps must be clean, action-oriented "
        "instructions.\n"
        "3. Don't include noise like 'Given CREATE EVENT FLOW Version 1' "
        "in steps.\n"
        "4. Rewrite Gherkin-style noise (e.g. 'Given [feature name]', "
        "'When [description]', 'Then [expected result]') into proper test "
        "steps.\n"
        "5. IMPORTANT: Do NOT use the current feature context to reject rows. "
        "This is PROJECT-WIDE memory. A row valid for another feature must "
        "keep is_testcase=true so it can be stored for future matching.\n"
        "6. Set is_testcase=false ONLY for genuine metadata rows: company/"
        "project name headers, environment details, document headers, section "
        "titles, version logs, column headers (e.g. 'Company Name', "
        "'Project Name', 'Document Type', 'Testing Type', 'Module Covered', "
        "'Environment', 'Serial No').\n"
        "7. Rows with endpoint lists, UI validations, edge cases, expected "
        "behaviour, or test-like verification intent are valid reusable QA "
        "memory even when they don't match the current feature.\n"
        "8. Tags: short, lowercase, relevant keywords. No version numbers/"
        "document types as tags.\n"
        "9. expected_result.summary: concise — one sentence describing what "
        "should happen.\n\n"
        f"CURRENT FEATURE (context only, NEVER an accept/reject filter):\n"
        f"Feature: {feature_name}\n"
        f"Description: {feature_description or 'N/A'}\n\n"
        f"IMPORT ROWS:\n{json.dumps(rows_data)[:6000]}\n\n"
        "FIELD REQUIREMENTS:\n"
        '- is_testcase: boolean\n'
        '- title: clean, descriptive\n'
        '- description: what this test verifies\n'
        '- intent: why this test matters\n'
        '- category: one of "api_tests", "ui_validations", "edge_cases", '
        '"business_tests", "e2e_tests"\n'
        '- test_suite: one of "Smoke", "Regression", "Edge", "Chaos"\n'
        '- priority: one of "High", "Medium", "Low"\n'
        '- endpoint: API path or null\n'
        '- method: HTTP verb or null\n'
        '- steps: array of {"content","expectedResult"}\n'
        '- expected_result: {"summary":"..."}\n'
        '- tags: array of short keywords\n\n'
        'Return JSON: {"polishedRows":[{"rowNumber":N,"is_testcase":bool,'
        '"title":"...","description":"...","intent":"...","category":"...",'
        '"test_suite":"...","priority":"...","endpoint":"..."|null,'
        '"method":"..."|null,"steps":[{"content":"...","expectedResult":"..."}],'
        '"expected_result":{"summary":"..."},"tags":["..."]}]}.'
    )
    try:
        return llm.chat_json(_POLISH_SYS, prompt, temperature=0.1,
                              max_tokens=4000, retries=0) or {}
    except Exception as e:  # noqa: BLE001
        return {"_error": str(e)}


def polish_all_rows(llm, rows: list[ParsedRow],
                     feature_name: str, feature_description: str,
                     batch_size: int = 8, max_workers: int = 6,
                     progress_fn=None) -> list[ParsedRow]:
    """Batched + concurrent polish pass. Mirrors Node's
    `polishImportedRowsWithAI` (batches of 8, concurrency up to 15).

    Drops rows where the LLM returns `is_testcase=false`. Applies polished
    fields onto a copy of each ParsedRow. Per-batch errors fall through —
    affected rows keep their parser-derived values.
    """
    if not rows:
        return rows
    from concurrent.futures import ThreadPoolExecutor, as_completed

    batches = [rows[i:i + batch_size] for i in range(0, len(rows), batch_size)]
    polished_by_row: dict[int, dict] = {}
    drop_row_numbers: set[int] = set()
    completed = 0

    def _do_batch(batch):
        return batch, polish_rows_batch(llm, batch, feature_name,
                                          feature_description)

    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = [pool.submit(_do_batch, b) for b in batches]
        for fut in as_completed(futures):
            batch, data = fut.result()
            for p in (data.get("polishedRows") or []):
                if not isinstance(p, dict) or "rowNumber" not in p:
                    continue
                rn = p.get("rowNumber")
                if not isinstance(rn, int):
                    try:
                        rn = int(rn)
                    except Exception:  # noqa: BLE001
                        continue
                if p.get("is_testcase") is False:
                    drop_row_numbers.add(rn)
                    continue
                polished_by_row[rn] = p
            completed += 1
            if progress_fn:
                progress_fn(completed, len(batches))

    out: list[ParsedRow] = []
    for r in rows:
        if r.row_number in drop_row_numbers:
            continue
        p = polished_by_row.get(r.row_number)
        if not p:
            out.append(r)
            continue
        # Apply polished fields back onto the row.
        if p.get("title"): r.title = str(p["title"])[:240]
        if p.get("description"): r.description = str(p["description"])[:1500]
        if p.get("intent"): r.intent = str(p["intent"])[:600]
        if p.get("category"): r.category = str(p["category"])
        if p.get("priority"):
            pr = str(p["priority"]).strip().lower()
            if pr.startswith("h"): r.priority = "High"
            elif pr.startswith("l"): r.priority = "Low"
            else: r.priority = "Mid"
        if p.get("endpoint") is not None:
            r.endpoint = str(p["endpoint"] or "")
        if p.get("method") is not None:
            r.method = str(p["method"] or "").upper()
        if p.get("test_suite"): r.suite = str(p["test_suite"])
        if p.get("tags"):
            r.tags = [str(t)[:40] for t in p["tags"] if t][:20]
        polished_steps = p.get("steps") or []
        if polished_steps:
            r.steps = []
            expecteds = []
            for s in polished_steps:
                if isinstance(s, dict):
                    if s.get("content"):
                        r.steps.append(str(s["content"])[:600])
                    if s.get("expectedResult"):
                        expecteds.append(str(s["expectedResult"])[:600])
                elif isinstance(s, str):
                    r.steps.append(s[:600])
            if expecteds and not r.expected_result:
                r.expected_result = " | ".join(expecteds)[:1500]
        if isinstance(p.get("expected_result"), dict):
            summary = (p["expected_result"].get("summary") or "").strip()
            if summary:
                r.expected_result = summary[:1500]
        out.append(r)
    return out


def build_xlsx_template() -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    headers = [c[0] for c in TEMPLATE_COLUMNS]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F497D")
    header_font = Font(bold=True, color="FFFFFF")
    for i, (name, width) in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=i)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 25
    ws.append(TEMPLATE_EXAMPLE)
    ws.row_dimensions[2].height = 65
    example_font = Font(italic=True, color="808080")
    for i in range(1, len(TEMPLATE_COLUMNS) + 1):
        c = ws.cell(row=2, column=i)
        c.font = example_font
        c.alignment = Alignment(wrap_text=True, vertical="top")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
